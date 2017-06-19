#!/usr/bin/python2

from __future__ import print_function
import bs4, mechanize, argparse, sqlite3, string, sys

from openpyxl import Workbook

br = None
options = None
sort = {}
results = set()
subjects = {}

class Result:
    def __init__(self, grade, benchno):
        p('\rCollecting %d ..' % benchno)
        # POST stuff ..
        br.select_form(nr=1)
        br['grade'] = [grade]
        br['beachno'] = str(benchno)
        br.submit()
        # Extracting marks ..
        bs = bs4.BeautifulSoup(br.response().read(), 'lxml')
        info = bs.find(attrs={'class': 'table_cc'})
        if info is None:
            br.back()
            raise ValueError()
        else:
            info = info.findAll('tr')

        self.name = str(info[0].contents[3].text).strip()
        if self.name == '':
            br.back()
            raise ValueError()
        n = self.name.split()[:3]
        if {'El': True, 'Al': True, 'Abdel': True, 'Abdul': True}.get(n[2], False):
            n = self.name.split()[:4]
        self.name = string.join(n)

        self.marks = {}
        for row in info[1:]:
            l = len(row.contents)
            if l < 5: continue
            line = str(row.contents[1].text).strip()
            subject = line[:-5].rstrip() if l == 5 else line.encode()
            try:
                mark = float(row.contents[3].text if l == 5 else row.contents[3].text[:-1])
            except:
                continue
            self.marks[subject] = mark
            subjects[subject] = int(line[-3 if subject != 'Total' else -4:-1]) if l == 5 else 100
        # Finalizing ..
        br.back()

class Writer:
    def __init__(self, form, name):
        self._write, self.name = {'text': (self._write_text, '.txt'), 'html': (self._write_html, '.html'), 'excel': (self._write_excel, '.xlsx'), 'sqlite': (self._write_sqlite, '.db')}.get(form)
        self.name = name + self.name
        self.form = form

    def write(self):
        self._write()
        print('Written in %s [%s]!' % (self.form, self.name))

    def _write_text(self):
        with open(self.name, 'w') as f:
            for subject in subjects:
                f.write('%s [%d]:\n' % (subject, subjects[subject]))
                for i, o in enumerate(sort[subject], 1):
                    f.write('\t%02d. %-40s\t--->\t%.2f%s\n' % (i, o.name, o.marks[subject], "!!" if o.marks[subject] > subjects[subject] else ''))
                    if i == options.tops: break

    def _write_html(self):
        with open(self.name, 'w') as f:
            f.write('<link rel="stylesheet" href="tablestyle.css" /><table><tr>')
            for subject in subjects: f.write('<th>%s [%d]</th>' % (subject, subjects[subject]))
            f.write('</tr>')
            for y in range(0, options.tops):
                f.write('<tr>')
                for subject in subjects:
                    try:
                        o = sort[subject][y]
                    except IndexError:
                        f.write('<td></td>')
                        continue
                    f.write('<td><span class="rank">%02d</span>. <span class="name">%s</span><span class="mark"> %.2f%s</span></td>' % (y+1, o.name, o.marks[subject], "!!" if o.marks[subject] > subjects[subject] else ''))
                f.write('</tr>')
            f.write('</table>')

    def _write_excel(self):
        wb = Workbook()
        wsl = wb.active
        x = 1
        for subject in sort:
            wsl.merge_cells(start_row=1, end_row=1, start_column=x, end_column=x+3)
            wsl.cell(row=1, column=x, value='%s [%s]' % (subject, subjects[subject]))
            wsl.cell(row=2, column=x, value='Rank')
            wsl.merge_cells(start_row=2, end_row=2, start_column=x+1, end_column=x+2)
            wsl.cell(row=2, column=x+1, value='Name')
            wsl.cell(row=2, column=x+3, value='Mark')
            for i, o in enumerate(sort[subject], 1):
                wsl.cell(row=i+2, column=x, value=i)
                wsl.merge_cells(start_row=i+2, end_row=i+2, start_column=x+1, end_column=x+2)
                wsl.cell(row=i+2, column=x+1, value=o.name)
                wsl.cell(row=i+2, column=x+3, value=o.marks[subject])
                if i == options.tops: break
            x += 4
        wb.save(self.name)

    def _write_sqlite(self):
        conn = sqlite3.connect(options.outfile + '.db')
        c = conn.cursor()
        c.execute('create table results (subject string, rank integer(3), name string, mark float(2))')
        for subject in sort:
            for i, o in enumerate(sort[subject], 1):
                c.execute('INSERT INTO results VALUES (?,?,?,?)', (subject, i, o.name, o.marks[subject],))
        conn.commit()
        c.close()
        conn.close()

def parse_args():
    parser = argparse.ArgumentParser(description="Ranks students' results", epilog='(C) 2017 -- Amr Ayman')
    parser.add_argument('-g', '--grade', required=True, choices=['J1', 'J2', 'J3', 'J4', 'J5', 'J6', 'M1', 'M2', 'M3', 'S1', 'S2'], help="Student's grade. e.g: J3, M2, ..")
    parser.add_argument('-o', '--outfile', required=True, help='Output filename')
    parser.add_argument('benchnos', nargs='+', type=int, help='Student bench numbers')
    parser.add_argument('-f', default=['html'], nargs='+', choices=['html', 'text', 'excel', 'sqlite'], help='Output file format. You can specify multiple, e.g: -f html excel ..', dest='fileformats')
    parser.add_argument( '--tops', default=10, type=int, help='How many tops ?')
    options = parser.parse_args()
    options.grade = {'J1': '1', 'J2': '2', 'J3': '3', 'J4': '4', 'J5': '5', 'J6': '6', 'M1': '7', 'M2': '8', 'M3': '9', 'S1': '10', 'S2': '11'}.get(options.grade)
    if options.tops > len(options.benchnos):
        options.tops = len(options.benchnos)
    options.outs = [ Writer(f, options.outfile) for f in options.fileformats ]
    return options

def sort_results(results):
    sorted_results = {}
    subs = []
    for subject in subjects:
        sorted_results[subject] = sorted([res for res in results if subject in res.marks], key=lambda result: result.marks[subject], reverse=True)
        # No activity, pe shit ..
        if sorted_results[subject][0].marks[subject] == sorted_results[subject][-1].marks[subject]:
            sorted_results.pop(subject)
            subs.append(subject)
    for s in subs: subjects.pop(s)
    return sorted_results

def p(p):
    print(p, end='')
    sys.stdout.flush()

if __name__ == '__main__':
    try:
        options = parse_args()
        br = mechanize.Browser()
        p('Connecting ...')
        br.open('http://new-sls.net/grades', timeout=20)
        for bench in options.benchnos:
            try:
                results.add(Result(options.grade, bench))
            except ValueError:
                print('Invalid Bench no: %s' % bench, file=sys.stderr)
        print('\rCollected All!        ')

        sort = sort_results(results)
        for writer in options.outs:
            try:
                writer.write()
            except Exception as e:
                print(e.message)
    except KeyboardInterrupt:
        print('\nExiting ..', file=sys.stderr)
        sys.exit(1)
